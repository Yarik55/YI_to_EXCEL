# ─── IMPORTS ─────────────────────────────────────────────────────────────────
# FIX 1: You had "openpyx1" (with a number 1) instead of "openpyxl" (with a letter L)
# This typo would crash the entire script immediately
import openpyxl

# FIX 2: Same typo — "openx1" and "openpyx1" changed to "openpyxl"
from openpyxl.styles import Font, PatternFill, Alignment

# "Give me the tool that converts column numbers to letters"
# e.g. get_column_letter(1) → "A",  get_column_letter(4) → "D"
from openpyxl.utils import get_column_letter

# "Give me the datetime tool for working with dates and times"
# e.g. datetime.now().strftime("%Y%m%d_%H%M") → "20260519_1430"
from datetime import datetime

# "Load our xero_api.py file and call it xero for short"
import xero_api as xero

# FIX 3: You had duplicate import lines scattered through the file
# All imports must be at the TOP only — removed the duplicates below


# ─── STYLE CONSTANTS ─────────────────────────────────────────────────────────
# These are defined once here and reused everywhere — change colour in one place,
# updates across all sheets automatically

# "Create a dark navy blue background fill"
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")

# "White bold text size 11 for column headers"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# "Large navy bold text size 14 for sheet titles"
TITLE_FONT  = Font(bold=True, size=14, color="1E3A5F")

# "Excel currency format — displays 1500 as $1,500.00"
MONEY_FMT   = '$#,##0.00'


# ─── HELPER FUNCTION 1 ───────────────────────────────────────────────────────
# "Define a function that styles a header row"
# Parameters: ws = worksheet, row_num = which row, num_cols = how many columns
def style_header_row(ws, row_num, num_cols):
    # "Loop through each column number from 1 to num_cols"
    for col in range(1, num_cols + 1):
        # "Grab the specific cell at this row and column position"
        cell = ws.cell(row=row_num, column=col)
        # "Apply the navy background to this cell"
        cell.fill = HEADER_FILL
        # "Apply the white bold font to this cell"
        cell.font = HEADER_FONT
        # "Centre the text horizontally in this cell"
        cell.alignment = Alignment(horizontal="center")


# ─── HELPER FUNCTION 2 ───────────────────────────────────────────────────────
# "Define a function that automatically sets each column width based on content"
def auto_width(ws):
    # "Loop through every column in the worksheet"
    for col in ws.columns:
        # "Find the longest text length across all cells in this column"
        # cell.value or "" means: if cell is empty use "" instead of None
        max_len = max(len(str(cell.value or "")) for cell in col)
        # "Set width to longest content + 4 padding, never wider than 50"
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ─── SHEET 1: INVOICES ───────────────────────────────────────────────────────
def write_invoices(wb):
    # "Create a new sheet called Invoices inside the workbook"
    ws = wb.create_sheet("Invoices")

    # "Put the title in cell A1 and style it"
    ws["A1"] = "Invoices"
    ws["A1"].font = TITLE_FONT

    # "Define column header names"
    headers = ["Invoice #", "Contact", "Date", "Due Date",
               "Status", "Amount Due", "Amount Paid"]

    # "Add headers as row 2"
    ws.append(headers)

    # "Style row 2 with navy background and white text"
    style_header_row(ws, 2, len(headers))

    # "Loop through every invoice Xero returns"
    for inv in xero.get_invoices():
        ws.append([
            # .get() is safe — returns None instead of crashing if field missing
            inv.get("InvoiceNumber"),
            # Contact is nested: {"Contact": {"Name": "Acme"}} — need two .get() calls
            inv.get("Contact", {}).get("Name"),
            # [:10] slices the date string to just "2026-05-19" (first 10 chars)
            inv.get("DateString", "")[:10],
            inv.get("DueDateString", "")[:10],
            inv.get("Status"),
            inv.get("AmountDue"),
            inv.get("AmountPaid"),
        ])

    # "Apply dollar format to columns 6 and 7 (Amount Due, Amount Paid)"
    for row in ws.iter_rows(min_row=3, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = MONEY_FMT

    # "Auto-size all columns"
    auto_width(ws)


# ─── SHEET 2: BANK TRANSACTIONS ──────────────────────────────────────────────
def write_bank_transactions(wb):
    ws = wb.create_sheet("Bank Transactions")

    ws["A1"] = "Bank Transactions"
    ws["A1"].font = TITLE_FONT

    headers = ["Date", "Contact", "Type", "Reference", "Amount", "Status"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    # "Loop through every bank transaction Xero returns"
    for txn in xero.get_bank_transactions():
        ws.append([
            txn.get("DateString", "")[:10],
            txn.get("Contact", {}).get("Name"),
            # Type values: RECEIVE (money in) or SPEND (money out)
            txn.get("Type"),
            txn.get("Reference"),
            # Bank transactions use "Total" not "AmountDue"
            txn.get("Total"),
            txn.get("Status"),
        ])

    # "Apply dollar format to column 5 only (Amount)"
    for row in ws.iter_rows(min_row=3, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = MONEY_FMT

    auto_width(ws)


# ─── SHEET 3: PROFIT & LOSS ──────────────────────────────────────────────────
# P&L comes back as a nested report structure (sections + rows)
# NOT a flat list like invoices — so we navigate it differently
def write_profit_and_loss(wb):
    ws = wb.create_sheet("Profit and Loss")

    ws["A1"] = "Profit and Loss"
    ws["A1"].font = TITLE_FONT

    headers = ["Account", "Amount"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    # "Get the full P&L report structure from Xero"
    report = xero.get_profit_and_loss()

    # "Loop through each top-level section in the report"
    for section in report.get("Rows", []):
        row_type = section.get("RowType")

        # "If this is a section group (e.g. Income, Expenses)"
        if row_type == "Section":
            title = section.get("Title", "")
            # "Only add the title if it has content (not empty)"
            if title:
                ws.append([title])
                # "Make the section title bold"
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

            # "Loop through individual account rows inside this section"
            for row in section.get("Rows", []):
                cells = row.get("Cells", [])
                # "Only process if row has at least 2 cells (name + amount)"
                if len(cells) >= 2:
                    ws.append([
                        cells[0].get("Value"),  # Account name
                        cells[1].get("Value"),  # Amount
                    ])

        # "If this is a total/summary row (e.g. Net Profit) — make it bold"
        elif row_type == "SummaryRow":
            cells = section.get("Cells", [])
            if len(cells) >= 2:
                ws.append([
                    cells[0].get("Value"),
                    cells[1].get("Value"),
                ])
                # "Bold both columns on summary rows so totals stand out"
                for col in range(1, 3):
                    ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    # "Apply dollar format to amount column"
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):
        for cell in row:
            try:
                # "Xero sends amounts as strings — convert to float first"
                cell.value = float(cell.value)
                cell.number_format = MONEY_FMT
            except (ValueError, TypeError):
                # "If conversion fails (e.g. cell contains text) skip quietly"
                pass

    auto_width(ws)


# ─── SHEET 4: BALANCE SHEET ──────────────────────────────────────────────────
# Same structure as P&L — sections, rows, summary rows
def write_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")

    ws["A1"] = "Balance Sheet"
    ws["A1"].font = TITLE_FONT

    headers = ["Account", "Amount"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    report = xero.get_balance_sheet()

    for section in report.get("Rows", []):
        row_type = section.get("RowType")

        if row_type == "Section":
            title = section.get("Title", "")
            if title:
                ws.append([title])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

            for row in section.get("Rows", []):
                cells = row.get("Cells", [])
                if len(cells) >= 2:
                    ws.append([
                        cells[0].get("Value"),
                        cells[1].get("Value"),
                    ])

        elif row_type == "SummaryRow":
            cells = section.get("Cells", [])
            if len(cells) >= 2:
                ws.append([
                    cells[0].get("Value"),
                    cells[1].get("Value"),
                ])
                for col in range(1, 3):
                    ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):
        for cell in row:
            try:
                cell.value = float(cell.value)
                cell.number_format = MONEY_FMT
            except (ValueError, TypeError):
                pass

    auto_width(ws)


# ─── MAIN RUNNER ─────────────────────────────────────────────────────────────
# FIX 4: "if __name__" was indented inside write_balance_sheet() by mistake
# It must be at the LEFT EDGE — zero indentation — not inside any function
# In Python, indentation = inside a block. Zero indent = top level of file.
if __name__ == "__main__":
    print("⏳ Creating workbook...")

    # "Create a brand new empty Excel workbook in memory"
    wb = openpyxl.Workbook()

    # "Remove the default blank sheet Excel always creates"
    wb.remove(wb.active)

    print("⏳ Fetching invoices from Xero...")
    write_invoices(wb)

    print("⏳ Fetching bank transactions from Xero...")
    write_bank_transactions(wb)

    print("⏳ Fetching Profit & Loss from Xero...")
    write_profit_and_loss(wb)

    print("⏳ Fetching Balance Sheet from Xero...")
    write_balance_sheet(wb)

    # "Build filename with today's date and time baked in"
    # e.g. xero_report_20260520_0825.xlsx
    filename = f"xero_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    # "Write the workbook from memory to disk as a real .xlsx file"
    wb.save(filename)

    print(f"✅ Done! Saved as {filename}")
